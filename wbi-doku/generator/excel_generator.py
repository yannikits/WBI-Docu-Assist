"""
Excel-Dokument-Generator
Kopiert die WBI-Vorlage (.xlsx) und befüllt sie mit Metadaten und Inhalt.
"""

import io
import re
from datetime import datetime
from copy import copy

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _safe_filename(title: str) -> str:
    return re.sub(r'[^\w\s\-äöüÄÖÜß]', '_', title).strip() + ".xlsx"


def _header_style():
    return {
        'font': Font(bold=True, color='FFFFFF'),
        'fill': PatternFill('solid', fgColor='1F3864'),
        'alignment': Alignment(horizontal='left', vertical='center'),
    }


def _apply_header(cell):
    s = _header_style()
    cell.font = s['font']
    cell.fill = s['fill']
    cell.alignment = s['alignment']


def generate_excel(data: dict, template_path: str):
    """
    Erstellt eine Excel-Datei basierend auf der WBI-Vorlage.
    Gibt (BytesIO, filename) zurück.
    """
    title = data.get('titleSubject', '') + ' - ' + data.get('titleTopic', '')
    title = title.strip(' -')
    template_id = data.get('template', 'netzwerk')
    chapters = data.get('chapters', [])
    refs = [r for r in data.get('refs', []) if r.get('num') or r.get('name')]
    markdown_content = data.get('markdownContent', '')

    wb = load_workbook(template_path)

    if template_id == 'netzwerk':
        _fill_netzwerk(wb, data, title)
    elif template_id == 'brief':
        _fill_brief(wb, data, title)
    else:
        _fill_generic(wb, data, title, chapters, refs, markdown_content)

    buf = io.BytesIO()
    wb.save(buf)
    return buf, _safe_filename(title)


def _fill_netzwerk(wb, data: dict, title: str):
    """Befüllt die Netzwerkdokumentation-Vorlage."""
    today = datetime.now().strftime('%d.%m.%Y')
    subject = data.get('titleSubject', '')

    # Info-Sheet
    if 'Info' in wb.sheetnames:
        ws = wb['Info']
        # Schreibe Werte in Spalte B (neben den Labels in A)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 'Kundenname':
                    ws.cell(row=cell.row, column=cell.column + 1, value=subject)
                elif cell.value == 'Netzwerkdoku':
                    ws.cell(row=cell.row, column=cell.column + 1, value=title)
                elif str(cell.value or '').startswith('Erstellt:'):
                    ws.cell(row=cell.row, column=cell.column + 1, value=today)


def _fill_brief(wb, data: dict, title: str):
    """Befüllt die Briefvorlage."""
    ws = wb.active
    today = datetime.now().strftime('%d.%m.%Y')
    # Suche nach leeren Feldern und fülle Datum ein
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or '').lower() in ('datum', 'date'):
                ws.cell(row=cell.row, column=cell.column + 1, value=today)
            if str(cell.value or '').lower() in ('betreff', 'subject'):
                ws.cell(row=cell.row, column=cell.column + 1, value=title)


def _fill_generic(wb, data: dict, title: str, chapters: list, refs: list, markdown: str):
    """Befüllt eine generische Excel-Vorlage mit der Dokumentstruktur."""
    ws = wb.active

    # Leere das Sheet
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    row_num = 1

    # Titel
    ws.cell(row=row_num, column=1, value=title)
    c = ws.cell(row=row_num, column=1)
    c.font = Font(bold=True, size=16, color='1F3864')
    ws.row_dimensions[row_num].height = 30
    row_num += 2

    # Datum
    ws.cell(row=row_num, column=1, value='Erstellt:')
    ws.cell(row=row_num, column=2, value=datetime.now().strftime('%d.%m.%Y'))
    row_num += 2

    # Mitgeltende Unterlagen
    if refs:
        ws.cell(row=row_num, column=1, value='Mitgeltende Unterlagen')
        _apply_header(ws.cell(row=row_num, column=1))
        _apply_header(ws.cell(row=row_num, column=2))
        ws.cell(row=row_num, column=1).value = 'DMS-Link / WBI-Nummer'
        ws.cell(row=row_num, column=2).value = 'Bezeichnung'
        row_num += 1
        for ref in refs:
            ws.cell(row=row_num, column=1, value=ref.get('num', ''))
            ws.cell(row=row_num, column=2, value=ref.get('name', ''))
            row_num += 1
        row_num += 1

    # Kapitelstruktur
    if markdown.strip():
        # Markdown parsen und als Zeilen eintragen
        for line in markdown.split('\n'):
            stripped = line.strip()
            if not stripped or stripped == '---':
                row_num += 1
                continue
            if stripped.startswith('# '):
                c = ws.cell(row=row_num, column=1, value=stripped[2:])
                c.font = Font(bold=True, size=14, color='1F3864')
            elif stripped.startswith('## '):
                c = ws.cell(row=row_num, column=1, value=re.sub(r'^\d+\.\s*', '', stripped[3:]))
                c.font = Font(bold=True, size=12, color='2E75B6')
            elif stripped.startswith('### '):
                c = ws.cell(row=row_num, column=1, value='    ' + re.sub(r'^[\d.]+\s*', '', stripped[4:]))
                c.font = Font(bold=True, size=11)
            elif stripped.startswith('|') and not re.match(r'^\|[-| :]+\|$', stripped):
                cells = [c.strip() for c in stripped.strip('|').split('|')]
                for ci, val in enumerate(cells, 1):
                    ws.cell(row=row_num, column=ci, value=val)
            elif not stripped.startswith(('>', '-', '*', '`', '#', '|')):
                ws.cell(row=row_num, column=1, value=stripped)
            row_num += 1
    else:
        valid_chapters = [c for c in chapters if c.get('name', '').strip()]
        for idx, ch in enumerate(valid_chapters, 1):
            c = ws.cell(row=row_num, column=1, value=f'{idx}. {ch["name"]}')
            c.font = Font(bold=True, size=12, color='1F3864')
            row_num += 1
            subs = [s for s in ch.get('subs', []) if s.strip()]
            for si, sub in enumerate(subs, 1):
                ws.cell(row=row_num, column=1, value=f'    {idx}.{si} {sub}')
                ws.cell(row=row_num, column=2, value='')
                row_num += 1
            row_num += 1

    # Spaltenbreiten anpassen
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 50
